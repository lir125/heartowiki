# -*- coding: utf-8 -*-
"""두근두근타운 도감 정보.xlsx 의 시트별 컬럼명(첫 행) 확인."""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)


def main():
    # 같은 폴더 또는 상위 폴더에서 xlsx 찾기
    candidates = [
        Path(__file__).parent / "두근두근타운 도감 정보.xlsx",
        Path(__file__).parent.parent / "두근두근타운 도감 정보.xlsx",
        Path.home() / "Desktop" / "두근두근타운 도감 정보.xlsx",
        Path.home() / "Documents" / "두근두근타운 도감 정보.xlsx",
    ]
    if len(sys.argv) > 1:
        candidates.insert(0, Path(sys.argv[1]).resolve())

    path = None
    for p in candidates:
        if p.exists():
            path = p
            break

    if not path:
        print("파일을 찾을 수 없습니다. 경로를 인자로 주거나 아래 위치에 두세요:")
        for p in candidates:
            print(" ", p)
        print("\n사용법: python check_xlsx_columns.py [파일경로]")
        sys.exit(1)

    out_path = path.with_suffix(".컬럼확인.txt")
    lines = [f"파일: {path}\n"]

    wb = load_workbook(path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0]) if rows else []
        headers_str = [str(h).strip() if h is not None else "" for h in headers]
        lines.append(f"[시트: {sheet_name}]")
        lines.append(f"  컬럼 수: {len(headers_str)}")
        lines.append(f"  컬럼명: {headers_str}")
        lines.append("")

    wb.close()
    text = "\n".join(lines)
    print(text)
    out_path.write_text(text, encoding="utf-8")
    print(f"결과 저장: {out_path}")


if __name__ == "__main__":
    main()
