# -*- coding: utf-8 -*-
"""
두근두근타운 도감 - 데스크톱 앱
실행 시 C:\\Users\\<사용자>\\Documents\\Heartowiki\\data 폴더를 사용하며,
도감 데이터·수집 정보·설정(색상 등)을 모두 해당 폴더에 JSON으로 저장합니다.
데이터·앱 업데이트는 구글 드라이브 또는 GitHub Releases로 가능합니다.
"""

import io
import json
import os
import stat
import subprocess
import sys
import time
import webbrowser
from pathlib import Path

import requests
import webview

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# 앱 버전 (앱 업데이트 확인 시 비교용)
APP_VERSION = "1.0.2"

# 리소스 경로 (exe 빌드 시)
def get_resource_dir():
    return Path(getattr(sys, "_MEIPASS", Path(__file__).parent.resolve()))

# 앱 데이터 폴더: C:\Users\<사용자>\Documents\Heartowiki\data
def get_data_dir() -> Path:
    data_dir = Path.home() / "Documents" / "Heartowiki" / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    # 최초 실행 시 config.json 없으면 기본 파일 생성 (GitHub에서 자동으로 도감 데이터 가져옴)
    config_file = data_dir / "config.json"
    if not config_file.exists():
        default_config = {
            "github_repo": "lir125/heartowiki",
            "github_data_branch": "main",
            "update_source": "github",
            "update_info_file_id": "",
            "github_update_path": "app_version.json",
        }
        config_file.write_text(json.dumps(default_config, ensure_ascii=False, indent=2), encoding="utf-8")
    return data_dir

RESOURCE_DIR = get_resource_dir()
DATA_DIR = get_data_dir()

CONFIG_PATH = DATA_DIR / "config.json"
COLLECTION_PATH = DATA_DIR / "collection.json"
SETTINGS_PATH = DATA_DIR / "settings.json"
CACHE_PATH = DATA_DIR / "cache.json"

_cached_base = None
_cached_user = None
_last_data_error = ""  # 데이터 로드 실패 시 사용자에게 표시할 메시지


def load_config() -> dict:
    """데이터 폴더의 config.json 로드."""
    default = {
        "github_repo": "lir125/heartowiki",
        "github_data_branch": "main",
        "update_source": "github",
        "update_info_file_id": "",
        "github_update_path": "app_version.json",
    }
    if not CONFIG_PATH.exists():
        return default
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            config = {**default, **data}
        if not (config.get("github_repo") or "").strip():
            config["github_repo"] = default["github_repo"]
        # 예전 config에 github_update_path가 비어 있으면 app_version.json 사용
        if not (config.get("github_update_path") or "").strip():
            config["github_update_path"] = default["github_update_path"]
        return config
    except Exception:
        return default


def _fetch_data_from_github(repo: str, branch: str = "main", path: str = "heartowiki.xlsx") -> dict:
    """GitHub 저장소에서 heartowiki.xlsx 다운로드 후 엑셀 파싱 → 도감 JSON 구조로 반환."""
    if not repo or "/" not in repo:
        raise ValueError("config.json에 github_repo(예: owner/repo)를 넣어 주세요.")
    repo = repo.strip()
    branch = (branch or "main").strip()
    path = (path or "heartowiki.xlsx").strip().lstrip("/")
    # 캐시 무효화: 새로고침 시 GitHub/CDN 캐시를 피해 최신 xlsx 반영
    url = f"https://raw.githubusercontent.com/{repo}/{branch}/{path}?t={int(time.time())}"
    headers = {"User-Agent": "Heartowiki/1.0", "Cache-Control": "no-cache", "Pragma": "no-cache"}
    r = requests.get(url, timeout=30, headers=headers)
    if r.status_code == 404:
        raise ValueError(
            f"GitHub에서 파일을 찾을 수 없습니다: {path}\n"
            f"저장소 {repo}의 {branch} 브랜치에 해당 파일이 있는지 확인해 주세요."
        )
    r.raise_for_status()

    # heartowiki.xlsx 다운로드 → 엑셀 파싱 → 도감 JSON 구조로 반환
    result = _xlsx_to_creatures_data(r.content)
    if "data_version" not in result:
        result["data_version"] = "1.0.1"
    return result


def _get_github_data_version(repo: str, branch: str = "main", path: str = "creatures_data.json") -> str:
    """GitHub에 올라온 도감 데이터의 data_version 값만 조회 (업데이트 여부 확인용). xlsx는 버전 비교 생략."""
    if not repo or "/" not in repo:
        return ""
    branch = (branch or "main").strip()
    path = (path or "creatures_data.json").strip().lstrip("/")
    if path.lower().endswith(".xlsx"):
        return ""  # xlsx는 버전 필드 없음, 업데이트 알림 생략
    url = f"https://raw.githubusercontent.com/{repo}/{branch}/{path}"
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()
        return str(data.get("data_version", "")).strip()
    except Exception:
        return ""


def _version_tuple(v: str):
    """버전 문자열을 (숫자, ...) 튜플로. '1.0.2' -> (1, 0, 2)."""
    try:
        v = (v or "").strip().lstrip("v")
        return tuple(int(x) for x in v.split(".") if x.isdigit())
    except Exception:
        return (0,)


def _check_update_github_file(repo: str, branch: str, path: str) -> dict:
    """GitHub 저장소의 버전 파일(JSON)에서 버전 확인. path 예: app_version.json."""
    if not repo or "/" not in repo:
        return {"hasUpdate": False}
    branch = (branch or "main").strip()
    path = (path or "app_version.json").strip().lstrip("/")
    url = f"https://raw.githubusercontent.com/{repo.strip()}/{branch}/{path}"
    try:
        r = requests.get(url, timeout=15, headers={"User-Agent": "Heartowiki/1.0"})
        r.raise_for_status()
        data = r.json()
    except Exception:
        return {"hasUpdate": False}

    remote_version = (data.get("version") or "").strip().lstrip("v")
    if not remote_version:
        return {"hasUpdate": False}
    if _version_tuple(remote_version) <= _version_tuple(APP_VERSION):
        return {"hasUpdate": False}

    download_url = (data.get("download_url") or "").strip()
    if not download_url:
        return {"hasUpdate": False}

    message = (data.get("message") or "").strip() or "새 버전이 있습니다."
    return {
        "hasUpdate": True,
        "version": remote_version,
        "message": message,
        "download_url": download_url,
        "exe_file_id": "",
    }


def _check_update_github(repo: str, config: dict) -> dict:
    """GitHub 업데이트 확인. github_update_path가 있으면 파일에서, 없으면 Releases에서 확인."""
    path = (config.get("github_update_path") or "app_version.json").strip() or "app_version.json"
    if path:
        branch = (config.get("github_data_branch") or "main").strip()
        return _check_update_github_file(repo, branch, path)

    if not repo or "/" not in repo:
        return {"hasUpdate": False}
    try:
        url = f"https://api.github.com/repos/{repo.strip()}/releases/latest"
        r = requests.get(url, timeout=15, headers={"Accept": "application/vnd.github.v3+json"})
        r.raise_for_status()
        data = r.json()
    except Exception:
        return {"hasUpdate": False}

    tag = (data.get("tag_name") or "").strip().lstrip("v")
    if not tag:
        return {"hasUpdate": False}
    if _version_tuple(tag) <= _version_tuple(APP_VERSION):
        return {"hasUpdate": False}

    assets = data.get("assets") or []
    download_url = ""
    for a in assets:
        if a.get("content_type") == "application/x-msdownload" or (a.get("name") or "").endswith(".exe"):
            download_url = (a.get("browser_download_url") or "").strip()
            break
    if not download_url and assets:
        download_url = (assets[0].get("browser_download_url") or "").strip()
    if not download_url:
        return {"hasUpdate": False}

    body = (data.get("body") or "").strip() or "새 버전이 있습니다."
    return {
        "hasUpdate": True,
        "version": tag,
        "message": body,
        "download_url": download_url,
        "exe_file_id": "",
    }


def _check_update_google_drive(file_id: str) -> dict:
    """구글 드라이브 앱 업데이트 정보 JSON으로 확인."""
    if not file_id:
        return {"hasUpdate": False}
    try:
        url = f"https://drive.google.com/uc?export=download&id={file_id}"
        session = requests.Session()
        response = session.get(url, timeout=15)
        for key, value in response.cookies.items():
            if key.startswith("download_warning"):
                response = session.get(url, params={"confirm": value}, timeout=15)
                break
        response.raise_for_status()
        data = response.json()
    except Exception:
        return {"hasUpdate": False}

    remote_version = (data.get("version") or "").strip()
    exe_file_id = (data.get("exe_file_id") or "").strip()
    message = (data.get("message") or "").strip() or "새 버전이 있습니다."
    if not remote_version or not exe_file_id:
        return {"hasUpdate": False}
    if _version_tuple(remote_version) <= _version_tuple(APP_VERSION):
        return {"hasUpdate": False}
    return {
        "hasUpdate": True,
        "version": remote_version,
        "message": message,
        "download_url": "",
        "exe_file_id": exe_file_id,
    }


def check_app_update() -> dict:
    """설정에 따라 구글 드라이브 또는 GitHub에서 업데이트 정보 확인."""
    config = load_config()
    source = (config.get("update_source") or "google_drive").strip().lower()
    if source == "github":
        repo = (config.get("github_repo") or "").strip()
        return _check_update_github(repo, config)
    file_id = (config.get("update_info_file_id") or "").strip()
    return _check_update_google_drive(file_id)


def _download_exe_to_path(download_url: str, save_path: Path) -> bool:
    """URL에서 exe 다운로드하여 save_path에 저장. 성공 시 True."""
    try:
        r = requests.get(download_url, timeout=120, stream=True)
        r.raise_for_status()
        with open(save_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except Exception:
        return False


def _download_exe_from_drive(file_id: str, save_path: Path) -> bool:
    """구글 드라이브에서 exe 다운로드하여 save_path에 저장."""
    try:
        url = f"https://drive.google.com/uc?export=download&id={file_id}"
        session = requests.Session()
        response = session.get(url, timeout=120, stream=True)
        for key, value in response.cookies.items():
            if key.startswith("download_warning"):
                response = session.get(url, params={"confirm": value}, timeout=120, stream=True)
                break
        response.raise_for_status()
        with open(save_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except Exception:
        return False


def apply_update(download_url: str = "", drive_file_id: str = "") -> dict:
    """
    새 exe를 다운로드한 뒤, 실행 중인 exe를 같은 경로·같은 파일 이름으로 교체하고 재시작.
    (사용자가 바탕화면에서 '두타위키.exe'로 실행했다면, 업데이트 후에도 같은 위치 같은 이름으로 유지)
    - download_url: GitHub 등 직접 다운로드 URL (우선)
    - drive_file_id: 구글 드라이브 파일 ID
    반환: { success: bool, error: str 또는 빈 문자열 }
    exe가 아닌 상태(스크립트 실행)에서는 success=False, 브라우저로 열기만 안내.
    """
    if not getattr(sys, "frozen", False):
        if download_url:
            webbrowser.open(download_url)
        elif drive_file_id:
            webbrowser.open(f"https://drive.google.com/uc?export=download&id={drive_file_id}")
        return {"success": False, "error": "exe로 실행 중일 때만 자동 업데이트가 가능합니다."}

    # 현재 실행 중인 exe 경로 = 원래 위치 + 원래 이름 (업데이트 후에도 동일하게 유지)
    exe_path = Path(sys.executable).resolve()
    exe_dir = exe_path.parent
    new_path = exe_dir / (exe_path.stem + "_new" + exe_path.suffix)

    if download_url:
        ok = _download_exe_to_path(download_url, new_path)
    elif drive_file_id:
        ok = _download_exe_from_drive(drive_file_id, new_path)
    else:
        return {"success": False, "error": "다운로드 경로가 없습니다."}

    if not ok or not new_path.exists():
        return {"success": False, "error": "다운로드에 실패했습니다."}

    # 배치: 프로세스 종료 → 원본을 _old로 변경 → 새 exe를 원래 이름으로 변경 → _old 삭제 (사용자가 직접 다시 실행)
    pid = os.getpid()
    exe_path_str = str(exe_path)
    new_path_str = str(new_path)
    backup_name = exe_path.stem + "_old" + exe_path.suffix
    original_name = exe_path.name
    batch_content = f'''@echo off
chcp 65001 >nul
timeout /t 3 /nobreak >nul
taskkill /PID {pid} /F >nul 2>&1
timeout /t 5 /nobreak >nul
ren "{exe_path_str}" "{backup_name}"
timeout /t 1 /nobreak >nul
ren "{new_path_str}" "{original_name}"
del "{backup_name}" 2>nul
(del "%~f0" 2>nul)
exit
'''
    batch_path = exe_dir / "_update_and_restart.bat"
    batch_path.write_text(batch_content, encoding="utf-8")

    CREATE_NO_WINDOW = 0x08000000 if sys.platform == "win32" else 0
    subprocess.Popen(
        ["cmd", "/c", str(batch_path)],
        cwd=str(exe_dir),
        creationflags=CREATE_NO_WINDOW,
        shell=False,
    )
    return {"success": True, "error": ""}


def exit_app() -> None:
    """앱 종료 (업데이트 적용 후 호출)."""
    sys.exit(0)


def open_download_url(url: str) -> None:
    """기본 브라우저에서 URL 열기."""
    webbrowser.open(url)


def _val(row: tuple, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _xlsx_to_creatures_data(raw: bytes) -> dict:
    """엑셀 바이트를 도감 형식 { 어류, 곤충, 조류, 요리 [, data_version ] } 로 변환.
    시트: 도감 정보(도감 버전, 마지막 업데이트), 어류 관찰(이름,레벨,위치,크기,가격,시간대,날씨,비고),
    새/곤충(이름,레벨,위치,세부위치,시간대,날씨), 미식 라이프(이름,레벨,재료,레시피,가격,비고).
    """
    if load_workbook is None:
        raise ValueError("엑셀 파일을 읽으려면 openpyxl 패키지가 필요합니다.")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    result = {"어류": [], "곤충": [], "조류": [], "요리": []}

    # 도감 정보: 도감 버전, 마지막 업데이트 (있으면 data_version 설정)
    if "도감 정보" in wb.sheetnames:
        ws_info = wb["도감 정보"]
        rows_info = list(ws_info.iter_rows(values_only=True))
        if rows_info:
            first_info = [str(c).strip() if c is not None else "" for c in rows_info[0]]
            i_ver = -1
            for i, c in enumerate(first_info):
                if (c or "").strip() == "도감 버전":
                    i_ver = i
                    break
            if i_ver >= 0 and len(rows_info) > 1:
                ver_val = rows_info[1][i_ver] if i_ver < len(rows_info[1]) else None
                if ver_val is not None and str(ver_val).strip():
                    result["data_version"] = str(ver_val).strip()

    def col_index(first_row: list, name: str) -> int:
        for i, c in enumerate(first_row):
            if (c or "").strip() == name:
                return i
        return -1

    def col_index_any(first_row: list, names: list) -> int:
        for n in names:
            i = col_index(first_row, n)
            if i >= 0:
                return i
        return -1

    # 어류 관찰: 이름/명칭, 레벨, 위치/지역, 크기, 가격, 시간대, 날씨, 비고
    for sheet_candidate in ("어류 관찰", "어류"):
        if sheet_candidate not in wb.sheetnames:
            continue
        ws = wb[sheet_candidate]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            break
        first = [str(c).strip() if c is not None else "" for c in rows[0]]
        i_name = col_index_any(first, ["이름", "명칭"])
        if i_name < 0:
            break
        i_level = col_index(first, "레벨")
        i_loc = col_index_any(first, ["위치", "지역"])
        i_size = col_index(first, "크기")
        i_price = col_index(first, "가격")
        i_time = col_index(first, "시간대")
        i_weather = col_index(first, "날씨")
        i_note = col_index(first, "비고")
        for row in rows[1:]:
            name = _val(row, i_name)
            if not name:
                continue
            result["어류"].append({
                "명칭": name,
                "지역": _val(row, i_loc),
                "세부지역": "",
                "레벨": _val(row, i_level),
                "날씨영향": _val(row, i_weather),
                "이미지": "",
                "크기": _val(row, i_size),
                "가격": _val(row, i_price),
                "시간대": _val(row, i_time),
                "비고": _val(row, i_note),
            })
        break

    # 새 관찰 일지 / 곤충 이야기: 이름, 레벨, 위치, 세부위치, 시간대, 날씨
    category_sheets = [("조류", ("새 관찰 일지", "조류", "새")), ("곤충", ("곤충 이야기", "곤충"))]
    for category, sheet_names in category_sheets:
        ws = None
        for sn in sheet_names:
            if sn in wb.sheetnames:
                ws = wb[sn]
                break
        if ws is None:
            for name in wb.sheetnames:
                if category in (name or ""):
                    ws = wb[name]
                    break
        if ws is None:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        first = [str(c).strip() if c is not None else "" for c in rows[0]]
        i_name = col_index_any(first, ["이름", "명칭"])
        if i_name < 0:
            continue
        i_level = col_index(first, "레벨")
        i_loc = col_index_any(first, ["위치", "지역"])
        i_sub = col_index(first, "세부위치")
        i_time = col_index(first, "시간대")
        i_weather = col_index_any(first, ["날씨", "날씨영향"])
        i_img = col_index(first, "이미지")
        for row in rows[1:]:
            name = _val(row, i_name)
            if not name:
                continue
            result[category].append({
                "명칭": name,
                "지역": _val(row, i_loc),
                "세부지역": _val(row, i_sub),
                "레벨": _val(row, i_level),
                "날씨영향": _val(row, i_weather),
                "이미지": _val(row, i_img) if i_img >= 0 else "",
                "시간대": _val(row, i_time),
                "비고": "",
            })

    # 미식 라이프: 이름, 레벨, 재료, 레시피, 가격, 비고
    if "미식 라이프" in wb.sheetnames:
        ws = wb["미식 라이프"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            first = [str(c).strip() if c is not None else "" for c in rows[0]]
            i_name = col_index(first, "이름")
            i_level = col_index(first, "레벨")
            i_ing = col_index(first, "재료")
            i_recipe = col_index(first, "레시피")
            i_price = col_index(first, "가격")
            i_note = col_index(first, "비고")
            if i_name >= 0:
                for row in rows[1:]:
                    name = _val(row, i_name)
                    if not name:
                        continue
                    result["요리"].append({
                        "명칭": name,
                        "레벨": _val(row, i_level) if i_level >= 0 else "",
                        "재료": _val(row, i_ing),
                        "레시피": _val(row, i_recipe),
                        "가격": _val(row, i_price),
                        "비고": _val(row, i_note),
                    })

    wb.close()
    return result


def _fetch_opensheet(spreadsheet_id: str) -> dict:
    """opensheet.elk.sh API로 공유된 Google Sheets를 JSON으로 읽기. 로그인 불필요.
    시트는 '링크가 있는 모든 사용자(보기)'로 공유되어 있어야 함.
    시트 순서: 1=도감정보, 2=어류 관찰, 3=새 관찰 일지, 4=곤충 이야기, 5=미식 라이프
    """
    base = "https://opensheet.elk.sh"
    result = {"어류": [], "곤충": [], "조류": [], "요리": []}

    def row_get(row: dict, key: str) -> str:
        v = row.get(key)
        return str(v).strip() if v is not None else ""

    # 시트 2 = 어류 관찰 (이름, 레벨, 위치, 크기, 가격, 시간대, 날씨, 비고)
    try:
        r = requests.get(f"{base}/{spreadsheet_id}/2", timeout=15)
        r.raise_for_status()
        rows = r.json()
        for row in (rows or []):
            name = row_get(row, "이름")
            if not name:
                continue
            result["어류"].append({
                "명칭": name,
                "지역": row_get(row, "위치"),
                "세부지역": "",
                "레벨": row_get(row, "레벨"),
                "날씨영향": row_get(row, "날씨"),
                "이미지": "",
                "크기": row_get(row, "크기"),
                "가격": row_get(row, "가격"),
                "시간대": row_get(row, "시간대"),
                "비고": row_get(row, "비고"),
            })
    except Exception:
        pass

    # 시트 3 = 새 관찰 일지, 4 = 곤충 이야기 (이름, 레벨, 위치, 세부위치, 시간대, 날씨)
    for sheet_num, category in [(3, "조류"), (4, "곤충")]:
        try:
            r = requests.get(f"{base}/{spreadsheet_id}/{sheet_num}", timeout=15)
            r.raise_for_status()
            rows = r.json()
            for row in (rows or []):
                name = row_get(row, "이름")
                if not name:
                    continue
                result[category].append({
                    "명칭": name,
                    "지역": row_get(row, "위치"),
                    "세부지역": row_get(row, "세부위치"),
                    "레벨": row_get(row, "레벨"),
                    "날씨영향": row_get(row, "날씨"),
                    "이미지": "",
                    "시간대": row_get(row, "시간대"),
                    "비고": "",
                })
        except Exception:
            pass

    # 시트 5 = 미식 라이프 (이름, 레벨, 재료, 레시피, 가격, 비고)
    try:
        r = requests.get(f"{base}/{spreadsheet_id}/5", timeout=15)
        r.raise_for_status()
        rows = r.json()
        for row in (rows or []):
            name = row_get(row, "이름")
            if not name:
                continue
            result["요리"].append({
                "명칭": name,
                "레벨": row_get(row, "레벨"),
                "재료": row_get(row, "재료"),
                "레시피": row_get(row, "레시피"),
                "가격": row_get(row, "가격"),
                "비고": row_get(row, "비고"),
            })
    except Exception:
        pass

    total = len(result["어류"]) + len(result["곤충"]) + len(result["조류"]) + len(result["요리"])
    if total == 0:
        raise ValueError("opensheet에서 데이터를 가져오지 못했습니다. 시트를 '링크가 있는 모든 사용자(보기)'로 공유했는지 확인하세요.")
    return result


def _download_google_sheets_xlsx(spreadsheet_id: str) -> bytes:
    """Google Sheets를 xlsx로 내보내기 URL로 다운로드. 시트가 '링크가 있는 모든 사용자'로 공유되어 있어야 함."""
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    r = requests.get(url, timeout=30, allow_redirects=True)
    r.raise_for_status()
    raw = r.content
    if raw[:2] != b"PK":
        raise ValueError("Google Sheets 응답이 엑셀 형식이 아닙니다. 시트를 '링크가 있는 모든 사용자(보기 권한)'로 공유했는지 확인하세요.")
    return raw


def download_from_google_drive(file_id: str) -> dict:
    """구글 드라이브 파일 또는 Google Sheets(스프레드시트)에서 도감 데이터 다운로드. JSON/엑셀 지원."""
    if not file_id or file_id == "YOUR_GOOGLE_DRIVE_FILE_ID":
        return {"어류": [], "곤충": [], "조류": [], "요리": []}

    # 스프레드시트 ID면 opensheet를 먼저 시도 (Drive URL은 시트에 대해 403 낼 수 있음)
    try:
        data = _fetch_opensheet(file_id)
        if data and (len(data.get("어류", [])) + len(data.get("곤충", [])) + len(data.get("조류", [])) + len(data.get("요리", []))) > 0:
            return data
    except Exception:
        pass

    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    session = requests.Session()
    try:
        response = session.get(url, timeout=30)
    except Exception as e:
        raise ValueError("네트워크 연결을 확인해 주세요. 스프레드시트 링크라면 시트를 '링크가 있는 모든 사용자(보기)'로 공유했는지 확인하세요. " + str(e))
    for key, value in response.cookies.items():
        if key.startswith("download_warning"):
            response = session.get(url, params={"confirm": value}, timeout=30)
            break
    response.raise_for_status()
    raw = response.content
    if len(raw) < 100 and (raw[:2] != b"PK" and (not raw.strip() or b"<" in raw[:200])):
        raise ValueError("Drive에서 파일을 받지 못했습니다(403 등). 스프레드시트 링크라면 시트를 '링크가 있는 모든 사용자(보기)'로 공유했는지 확인하세요.")

    # JSON 시도
    try:
        data = json.loads(raw.decode("utf-8"))
        if isinstance(data, dict) and "어류" in data:
            return data
    except Exception:
        pass

    # 엑셀 시도 (xlsx 매직 바이트 PK)
    if raw[:2] == b"PK":
        return _xlsx_to_creatures_data(raw)

    # Drive 파일이 아닌 경우: Google Sheets export 시도 후, 실패하면 opensheet.elk.sh 로 시트 ID 직접 읽기
    sheets_err = None
    try:
        raw = _download_google_sheets_xlsx(file_id)
        return _xlsx_to_creatures_data(raw)
    except Exception as e:
        sheets_err = str(e)

    # 스프레드시트 링크(docs.google.com/spreadsheets/d/ID/edit)일 때: opensheet API로 로그인 없이 JSON 조회
    try:
        return _fetch_opensheet(file_id)
    except Exception as e2:
        pass

    err_extra = (" 원인: " + sheets_err) if sheets_err else ""
    raise ValueError(
        "데이터를 받지 못했습니다. "
        "1) 스프레드시트 링크라면 시트를 [공유] → [링크가 있는 모든 사용자(보기)]로 설정했는지 확인하세요. "
        "2) 또는 [파일] → [다운로드] → [Excel]로 저장한 뒤 그 파일을 드라이브에 올리고, 그 파일의 링크에서 파일 ID를 넣어 보세요." + err_extra
    )


def load_collection() -> dict:
    """collection.json 로드 (수집 별, 사용자 추가 생물). 가격 별은 저장하지 않음."""
    default = {"stars": {}, "userCreatures": {"어류": [], "곤충": [], "조류": [], "요리": []}}
    if not COLLECTION_PATH.exists():
        return default
    try:
        with open(COLLECTION_PATH, "r", encoding="utf-8") as f:
            return {**default, **json.load(f)}
    except Exception:
        return default


def save_collection(data: dict) -> None:
    """collection.json 저장."""
    with open(COLLECTION_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_settings_file() -> dict:
    """settings.json 로드 (탭, 정렬, 색상 등)."""
    default = {"currentTab": "어류", "sortBy": "level-asc", "colors": {}}
    if not SETTINGS_PATH.exists():
        return default
    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            return {**default, **json.load(f)}
    except Exception:
        return default


def save_settings_file(data: dict) -> None:
    """settings.json 저장."""
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_user_data() -> dict:
    """collection + settings 합쳐서 반환 (UI용)."""
    collection = load_collection()
    settings = load_settings_file()
    return {
        "stars": collection.get("stars", {}),
        "userCreatures": collection.get("userCreatures", {"어류": [], "곤충": [], "조류": [], "요리": []}),
        "settings": settings,
    }


def get_base_data() -> dict:
    """도감 데이터: GitHub 또는 구글 드라이브에서 받거나 데이터 폴더 캐시 사용."""
    global _cached_base, _last_data_error
    if _cached_base is not None:
        return _cached_base

    config = load_config()
    _last_data_error = ""

    try:
        # 데이터는 GitHub의 heartowiki.xlsx만 사용 (다운로드 → xlsx 파싱 → JSON 구조로 캐시)
        repo = (config.get("github_repo") or "lir125/heartowiki").strip()
        branch = (config.get("github_data_branch") or "main").strip()
        path = "heartowiki.xlsx"  # 항상 저장소 루트의 heartowiki.xlsx
        _cached_base = _fetch_data_from_github(repo, branch, path)
        # 성공 시 데이터 폴더에 JSON으로 캐시 저장 (cache.json)
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(_cached_base, f, ensure_ascii=False, indent=2)
    except Exception as e:
        _last_data_error = str(e) or "알 수 없는 오류"
        if CACHE_PATH.exists():
            try:
                with open(CACHE_PATH, "r", encoding="utf-8") as f:
                    _cached_base = json.load(f)
            except Exception:
                _cached_base = {"어류": [], "곤충": [], "조류": [], "요리": []}
        else:
            _cached_base = {"어류": [], "곤충": [], "조류": [], "요리": []}
    return _cached_base


def get_app_data() -> dict:
    """UI에서 호출: 도감 데이터 + 수집정보/설정 한 번에 반환."""
    global _cached_user
    base = get_base_data()
    user = load_user_data()
    _cached_user = user
    data_version = base.get("data_version", "") if isinstance(base, dict) else ""
    return {
        "base": base,
        "user": user,
        "lastError": _last_data_error,
        "appVersion": APP_VERSION,
        "dataVersion": data_version,
    }


def save_user_data_from_app(stars=None, user_creatures=None, settings=None) -> None:
    """UI에서 호출: 수집정보·설정을 데이터 폴더 JSON으로 저장. 가격 별은 저장하지 않음."""
    global _cached_user

    if stars is not None or user_creatures is not None:
        collection = load_collection()
        if stars is not None:
            collection["stars"] = stars
        if user_creatures is not None:
            collection["userCreatures"] = user_creatures
        save_collection(collection)

    if settings is not None:
        save_settings_file(settings)

    _cached_user = load_user_data()


def refresh_data() -> dict:
    """UI에서 호출: 원격(GitHub/드라이브)에서 다시 받고 사용자 데이터와 함께 반환."""
    global _cached_base, _last_data_error
    _cached_base = None
    _last_data_error = ""
    return get_app_data()


def check_data_update() -> dict:
    """GitHub heartowiki.xlsx 기준 도감 데이터 업데이트 여부 확인 (xlsx는 버전 필드 없어 비움)."""
    config = load_config()
    repo = config.get("github_repo", "").strip()
    if not repo:
        return {"hasUpdate": False, "currentVersion": "", "latestVersion": ""}
    branch = config.get("github_data_branch") or "main"
    path = "heartowiki.xlsx"
    latest = _get_github_data_version(repo, branch, path)
    current = ""
    if CACHE_PATH.exists():
        try:
            with open(CACHE_PATH, "r", encoding="utf-8") as f:
                cached = json.load(f)
                current = str(cached.get("data_version", "")).strip()
        except Exception:
            pass
    if not latest:
        return {"hasUpdate": False, "currentVersion": current, "latestVersion": ""}
    has_update = bool(current != latest)
    return {"hasUpdate": has_update, "currentVersion": current, "latestVersion": latest}


class Api:
    def get_app_data(self):
        return get_app_data()

    def save_user_data(self, stars=None, user_creatures=None, settings=None):
        save_user_data_from_app(stars=stars, user_creatures=user_creatures, settings=settings)

    def refresh_data(self):
        return refresh_data()

    def check_app_update(self):
        return check_app_update()

    def check_data_update(self):
        return check_data_update()

    def apply_update(self, download_url="", drive_file_id=""):
        return apply_update(download_url=download_url, drive_file_id=drive_file_id)

    def exit_app(self):
        exit_app()

    def open_download_url(self, url):
        open_download_url(url)


def main():
    index_path = RESOURCE_DIR / "index.html"
    if not index_path.exists():
        if not getattr(sys, "frozen", False):
            print(f"index.html을 찾을 수 없습니다: {index_path}")
        return

    get_data_dir()
    get_base_data()

    window = webview.create_window(
        "두타위키",
        f"file:///{index_path.as_posix()}",
        width=1200,
        height=800,
        min_size=(800, 600),
        js_api=Api(),
    )
    webview.start(debug=False)


if __name__ == "__main__":
    main()
